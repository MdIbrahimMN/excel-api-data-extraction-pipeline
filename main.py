import pandas as pd
import requests
import time
import logging
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


# ============================================================
# 1. LOGGING CONFIGURATION
# ============================================================

logging.basicConfig(
    filename="pipeline.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

logging.info("========================================")
logging.info("Pipeline started")
logging.info("========================================")


# ============================================================
# 2. READ INPUT EXCEL
# ============================================================

try:

    df = pd.read_excel("input.xlsx")

    print("Excel file loaded successfully!")
    print(df)

    logging.info("Input Excel loaded successfully")

except Exception as e:

    print("Error reading Excel file:")
    print(e)

    logging.error("Error reading Excel: %s", e)

    exit()


# ============================================================
# 3. CHECK COLUMN
# ============================================================

if "keyword" not in df.columns:

    print("\nERROR: Column 'keyword' was not found.")

    print("Available columns:")
    print(df.columns.tolist())

    logging.error(
        "Column 'keyword' not found. Available columns: %s",
        df.columns.tolist()
    )

    exit()


# ============================================================
# 4. CLEAN INPUT DATA
# ============================================================

# Convert keyword values to strings
df["keyword"] = (
    df["keyword"]
    .astype(str)
    .str.strip()
)

# Remove empty values
df = df[df["keyword"] != ""]

# Remove duplicate keywords
df = df.drop_duplicates(
    subset=["keyword"]
)

# Reset index
df = df.reset_index(drop=True)

print("\nCleaned keywords:")
print(df)

logging.info(
    "Input cleaned. Total unique keywords: %d",
    len(df)
)


# ============================================================
# 5. API CONFIGURATION
# ============================================================

search_url = "https://en.wikipedia.org/w/api.php"

summary_base_url = (
    "https://en.wikipedia.org/api/rest_v1/page/summary/"
)

headers = {
    "User-Agent": "APIProject/1.0"
}


# ============================================================
# 6. STORE RESULTS
# ============================================================

all_results = []


# ============================================================
# 7. PROCESS EACH KEYWORD
# ============================================================

for keyword in df["keyword"]:

    print("\n========================================")
    print("Searching for:", keyword)
    print("========================================")

    logging.info(
        "Searching for keyword: %s",
        keyword
    )

    try:

        # ====================================================
        # 7.1 SEARCH API
        # ====================================================

        params = {
            "action": "query",
            "list": "search",
            "srsearch": keyword,
            "format": "json",
            "srlimit": 5
        }

        response = requests.get(
            search_url,
            params=params,
            headers=headers,
            timeout=10
        )

        print(
            "Search Status:",
            response.status_code
        )

        response.raise_for_status()

        data = response.json()

        search_results = (
            data
            .get("query", {})
            .get("search", [])
        )

        print(
            "Results found:",
            len(search_results)
        )

        logging.info(
            "Found %d results for %s",
            len(search_results),
            keyword
        )


        # ====================================================
        # 7.2 NO RESULTS
        # ====================================================

        if not search_results:

            print(
                "No results found for:",
                keyword
            )

            logging.warning(
                "No results found for %s",
                keyword
            )

            all_results.append({

                "Keyword": keyword,
                "Title": "",
                "Description": "",
                "Information": "No results found",
                "URL": "",
                "Source": "Wikipedia",
                "Collected_Date": datetime.now().strftime(
                    "%Y-%m-%d %H:%M:%S"
                ),
                "Status": "No Results"

            })

            continue


        # ====================================================
        # 7.3 PROCESS EACH SEARCH RESULT
        # ====================================================

        for result in search_results:

            page_title = result.get(
                "title",
                ""
            )

            if not page_title:
                continue

            print(
                "\nGetting:",
                page_title
            )

            logging.info(
                "Getting page: %s",
                page_title
            )


            try:

                # =================================================
                # PAGE SUMMARY API
                # =================================================

                summary_url = (
                    summary_base_url
                    + requests.utils.quote(page_title)
                )

                summary_response = requests.get(
                    summary_url,
                    headers=headers,
                    timeout=10
                )

                print(
                    "Summary Status:",
                    summary_response.status_code
                )

                summary_response.raise_for_status()

                page_data = summary_response.json()


                # =================================================
                # EXTRACT INFORMATION
                # =================================================

                title = page_data.get(
                    "title",
                    page_title
                )

                description = page_data.get(
                    "description",
                    ""
                )

                information = page_data.get(
                    "extract",
                    ""
                )

                page_url = (
                    page_data
                    .get("content_urls", {})
                    .get("desktop", {})
                    .get("page", "")
                )


                # =================================================
                # COLLECTION DATE
                # =================================================

                collected_date = datetime.now().strftime(
                    "%Y-%m-%d %H:%M:%S"
                )


                # =================================================
                # SAVE SUCCESSFUL RESULT
                # =================================================

                all_results.append({

                    "Keyword": keyword,

                    "Title": title,

                    "Description": description,

                    "Information": information,

                    "URL": page_url,

                    "Source": "Wikipedia",

                    "Collected_Date": collected_date,

                    "Status": "Success"

                })


                logging.info(
                    "Successfully collected: %s",
                    page_title
                )


            # =====================================================
            # PAGE TIMEOUT
            # =====================================================

            except requests.exceptions.Timeout:

                print(
                    "Timeout while getting:",
                    page_title
                )

                logging.error(
                    "Timeout while getting page: %s",
                    page_title
                )

                all_results.append({

                    "Keyword": keyword,

                    "Title": page_title,

                    "Description": "",

                    "Information": "",

                    "URL": "",

                    "Source": "Wikipedia",

                    "Collected_Date": datetime.now().strftime(
                        "%Y-%m-%d %H:%M:%S"
                    ),

                    "Status": "Page Timeout"

                })


            # =====================================================
            # PAGE REQUEST ERROR
            # =====================================================

            except requests.exceptions.RequestException as e:

                print(
                    "Error getting page:",
                    e
                )

                logging.error(
                    "Page error for %s: %s",
                    page_title,
                    e
                )

                # Use search snippet as backup
                snippet = result.get(
                    "snippet",
                    ""
                )

                all_results.append({

                    "Keyword": keyword,

                    "Title": page_title,

                    "Description": "",

                    "Information": snippet,

                    "URL": "",

                    "Source": "Wikipedia",

                    "Collected_Date": datetime.now().strftime(
                        "%Y-%m-%d %H:%M:%S"
                    ),

                    "Status": "Page Error"

                })


            # =====================================================
            # UNEXPECTED PAGE ERROR
            # =====================================================

            except Exception as e:

                print(
                    "Unexpected page error:",
                    e
                )

                logging.exception(
                    "Unexpected page error for %s",
                    page_title
                )

                all_results.append({

                    "Keyword": keyword,

                    "Title": page_title,

                    "Description": "",

                    "Information": "",

                    "URL": "",

                    "Source": "Wikipedia",

                    "Collected_Date": datetime.now().strftime(
                        "%Y-%m-%d %H:%M:%S"
                    ),

                    "Status": "Error"

                })


            # Small delay between requests
            time.sleep(0.5)


    # ========================================================
    # 7.4 SEARCH TIMEOUT
    # ========================================================

    except requests.exceptions.Timeout:

        print(
            "Request timed out for:",
            keyword
        )

        logging.error(
            "Request timeout for %s",
            keyword
        )

        all_results.append({

            "Keyword": keyword,

            "Title": "",

            "Description": "",

            "Information": "",

            "URL": "",

            "Source": "Wikipedia",

            "Collected_Date": datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),

            "Status": "Timeout"

        })


    # ========================================================
    # 7.5 API ERROR
    # ========================================================

    except requests.exceptions.RequestException as e:

        print(
            "API error for:",
            keyword
        )

        print(e)

        logging.error(
            "API error for %s: %s",
            keyword,
            e
        )

        all_results.append({

            "Keyword": keyword,

            "Title": "",

            "Description": "",

            "Information": "",

            "URL": "",

            "Source": "Wikipedia",

            "Collected_Date": datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),

            "Status": "API Error"

        })


    # ========================================================
    # 7.6 UNEXPECTED ERROR
    # ========================================================

    except Exception as e:

        print(
            "Unexpected error:",
            e
        )

        logging.exception(
            "Unexpected error for %s",
            keyword
        )

        all_results.append({

            "Keyword": keyword,

            "Title": "",

            "Description": "",

            "Information": "",

            "URL": "",

            "Source": "Wikipedia",

            "Collected_Date": datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),

            "Status": "Error"

        })


# ============================================================
# 8. CREATE OUTPUT DATAFRAME
# ============================================================

output_df = pd.DataFrame(
    all_results
)


# ============================================================
# 9. SAVE OUTPUT EXCEL
# ============================================================

output_file = "output.xlsx"


try:

    # --------------------------------------------------------
    # Save DataFrame
    # --------------------------------------------------------

    output_df.to_excel(
        output_file,
        index=False,
        engine="openpyxl"
    )


    # --------------------------------------------------------
    # Open workbook
    # --------------------------------------------------------

    workbook = load_workbook(
        output_file
    )

    worksheet = workbook.active


    # --------------------------------------------------------
    # 9.1 FORMAT HEADER
    # --------------------------------------------------------

    for cell in worksheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )


    # --------------------------------------------------------
    # 9.2 WRAP TEXT
    # --------------------------------------------------------

    for row in worksheet.iter_rows():

        for cell in row:

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )


    # --------------------------------------------------------
    # 9.3 COLUMN WIDTHS
    # --------------------------------------------------------

    worksheet.column_dimensions["A"].width = 15
    worksheet.column_dimensions["B"].width = 35
    worksheet.column_dimensions["C"].width = 35
    worksheet.column_dimensions["D"].width = 80
    worksheet.column_dimensions["E"].width = 60
    worksheet.column_dimensions["F"].width = 18
    worksheet.column_dimensions["G"].width = 22
    worksheet.column_dimensions["H"].width = 18


    # --------------------------------------------------------
    # 9.4 FREEZE HEADER
    # --------------------------------------------------------

    worksheet.freeze_panes = "A2"


    # --------------------------------------------------------
    # 9.5 ADD FILTER
    # --------------------------------------------------------

    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )


    # --------------------------------------------------------
    # 9.6 ADD EXCEL TABLE
    # --------------------------------------------------------

    if worksheet.max_row >= 2:

        table_ref = (
            f"A1:H{worksheet.max_row}"
        )

        table = Table(
            displayName="SearchResults",
            ref=table_ref
        )

        table_style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        table.tableStyleInfo = table_style

        worksheet.add_table(
            table
        )


    # --------------------------------------------------------
    # 9.7 SAVE FORMATTED WORKBOOK
    # --------------------------------------------------------

    workbook.save(
        output_file
    )


    # ========================================================
    # SUCCESS MESSAGE
    # ========================================================

    print("\n========================================")
    print("DATA COLLECTION COMPLETED")
    print("========================================")

    print(
        "Keywords processed:",
        len(df)
    )

    print(
        "Results collected:",
        len(output_df)
    )

    print(
        "Output file:",
        output_file
    )

    print(
        "Excel formatting completed!"
    )

    logging.info(
        "Output Excel created and formatted successfully"
    )


# ============================================================
# 10. PERMISSION ERROR
# ============================================================

except PermissionError:

    print("\nERROR:")

    print(
        "Please close output.xlsx and run the program again."
    )

    logging.error(
        "Permission denied while saving output.xlsx"
    )


# ============================================================
# 11. OTHER EXCEL ERROR
# ============================================================

except Exception as e:

    print(
        "Error saving or formatting Excel:",
        e
    )

    logging.exception(
        "Error saving output Excel"
    )


# ============================================================
# 12. FINISH
# ============================================================

logging.info(
    "Pipeline finished"
)

logging.info(
    "========================================"
)

print("\nPipeline finished successfully!")